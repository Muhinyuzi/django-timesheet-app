from django.shortcuts import render, redirect
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.db import IntegrityError
from django.forms import modelformset_factory
from decimal import Decimal
from django.db.models import Prefetch, Case, When, IntegerField
from datetime import timedelta
from datetime import date
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from django.contrib.auth.decorators import login_required

# Create your views here.

from .models import Employee, WeeklyTimesheet, DailyEntry
from .forms import WeeklyTimesheetForm, DailyEntryForm

weekday_order = Case(
    When(day="MON", then=1),
    When(day="TUE", then=2),
    When(day="WED", then=3),
    When(day="THU", then=4),
    When(day="FRI", then=5),
    When(day="SAT", then=6),
    When(day="SUN", then=7),
    output_field=IntegerField(),
)

def home(request):
    return render(request, "timesheet/home.html")


def employee_list(request):
    employees = Employee.objects.filter(is_active=True)
    return render(request, "timesheet/employee_list.html", {
        "employees": employees
    })


def timesheet_list(request):
    sort = request.GET.get("sort", "date")  # date par défaut

    timesheets = WeeklyTimesheet.objects.select_related("employee")

    if sort == "name":
        timesheets = timesheets.order_by("employee__name", "-week_start")
    else:  # tri par date (par défaut)
        timesheets = timesheets.order_by("-week_start", "employee__name")

    return render(request, "timesheet/timesheet_list.html", {
        "timesheets": timesheets,
        "sort": sort,
    })

@login_required
def timesheet_create(request):
    if request.method == "POST":
        form = WeeklyTimesheetForm(request.POST)
        if form.is_valid():
            try:
                ts = form.save()
                messages.success(request, "Feuille de temps créée avec succès.")
                # prochaine étape: rediriger vers detail page
                return redirect("timesheet:timesheet_detail", pk=ts.pk)
            except IntegrityError:
                form.add_error(None, "Une feuille existe déjà pour cet employé et cette semaine.")
    else:
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        form = WeeklyTimesheetForm(initial={"week_start": monday})

    return render(request, "timesheet/timesheet_form.html", {"form": form})

@login_required
def timesheet_detail(request, pk):
    timesheet = get_object_or_404(
        WeeklyTimesheet.objects.select_related("employee"),
        pk=pk
    )
    employee = timesheet.employee

    # S'assurer que Lun → Ven existent
    for day_code, _ in DailyEntry.Weekday.choices:
        DailyEntry.objects.get_or_create(timesheet=timesheet, day=day_code)

    queryset = (
        timesheet.entries.all()
        .annotate(day_order=weekday_order)
        .order_by("day_order")
    )

    DailyEntryFormSet = modelformset_factory(DailyEntry, form=DailyEntryForm, extra=0)

    if request.method == "POST":
        formset = DailyEntryFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Feuille mise à jour.")
            return redirect("timesheet:timesheet_detail", pk=timesheet.pk)
    else:
        formset = DailyEntryFormSet(queryset=queryset)

    # ✅ Totaux employé (toutes les semaines) — optimisé
    timesheets = (
        WeeklyTimesheet.objects
        .filter(employee=employee)
        .order_by("-week_start")
        .prefetch_related("entries")
    )

    total_hours = Decimal("0.00")
    total_regular = Decimal("0.00")
    total_banked = Decimal("0.00")

    for ts in timesheets:
        total_hours += ts.total_hours_decimal
        total_regular += ts.regular_hours
        total_banked += ts.banked_hours

    total_pay = (total_regular * employee.hourly_rate).quantize(Decimal("0.01"))

    return render(request, "timesheet/timesheet_detail.html", {
        "timesheet": timesheet,
        "formset": formset,
        "total_hours": total_hours,
        "total_banked": total_banked,
        "total_pay": total_pay,
    })

def payroll_summary(request):
    employees = Employee.objects.filter(is_active=True).prefetch_related(
        Prefetch("timesheets", queryset=WeeklyTimesheet.objects.all().order_by("-week_start"))
    )

    rows = []
    grand_total_hours = Decimal("0.00")
    grand_regular_hours = Decimal("0.00")
    grand_banked_hours = Decimal("0.00")
    grand_pay = Decimal("0.00")

    for emp in employees:
        total_hours = Decimal("0.00")
        regular_hours = Decimal("0.00")
        banked_hours = Decimal("0.00")

        for ts in emp.timesheets.all():
            total_hours += ts.total_hours_decimal
            regular_hours += ts.regular_hours
            banked_hours += ts.banked_hours

        pay_total = (regular_hours * emp.hourly_rate).quantize(Decimal("0.01"))

        grand_total_hours += total_hours
        grand_regular_hours += regular_hours
        grand_banked_hours += banked_hours
        grand_pay += pay_total

        rows.append({
            "employee": emp,
            "total_hours": total_hours,
            "regular_hours": regular_hours,
            "banked_hours": banked_hours,
            "hourly_rate": emp.hourly_rate,
            "pay_total": pay_total,
        })

    context = {
        "rows": rows,
        "grand_total_hours": grand_total_hours,
        "grand_regular_hours": grand_regular_hours,
        "grand_banked_hours": grand_banked_hours,
        "grand_pay": grand_pay.quantize(Decimal("0.01")),
    }
    return render(request, "timesheet/payroll_summary.html", context)

def export_timesheet_excel(request, pk):
    timesheet = get_object_or_404(WeeklyTimesheet, pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feuille de temps"

    # Titre
    ws["A1"] = "Employé"
    ws["B1"] = timesheet.employee.name

    ws["A2"] = "Semaine"
    ws["B2"] = f"{timesheet.week_start} → {timesheet.week_end}"

    # En-têtes
    headers = [
        "Jour",
        "Arrivée matin",
        "Départ dîner",
        "Retour dîner",
        "Arrivée soir",
        "Départ soir",
        "Total (h)",
    ]

    ws.append([])
    ws.append(headers)

    for cell in ws[4]:
        cell.font = Font(bold=True)

    # Données
    entries = timesheet.entries.all().annotate(day_order=weekday_order).order_by("day_order")

    for entry in entries:
        ws.append([
            entry.get_day_display(),
            entry.arrival_morning,
            entry.lunch_departure,
            entry.lunch_return,
            entry.arrival_evening,
            entry.departure_evening,
            entry.total_hours,
        ])

    ws.append([])
    ws.append(["Total semaine", "", "", "", "", "", timesheet.total_hours_decimal])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="timesheet_{timesheet.pk}.xlsx"'

    wb.save(response)
    return response